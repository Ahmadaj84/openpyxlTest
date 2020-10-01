import openpyxl as oxl
from openpyxl import load_workbook 

file_name = "C:\\Users\\ajifry\\Documents\\x.xlsx"
#load the workbook

wb = load_workbook(file_name)
#get list of sheets in the workbook

sheet_list = wb.sheetnames
#loop over each sheet

for sheet in sheet_list:
    s = wb[sheet]
    x = len(s['A'])
    print (s.title)
    #loop over each row in the sheet

    while x > 3:
        cell = 'C'+ str(x)
        #check if its a merge cell or not

        if type(s[cell]).__name__ == 'MergedCell':
            cell = 'B'+ str(x)
            #check the content of the cell
        if "مدير" not in s[cell].value:
            print (s[cell].value)
        x-=1
